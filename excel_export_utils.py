from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
try:
    from tqdm import tqdm
except ImportError:
    def tqdm(iterable=None, **_kwargs):
        return iterable


logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


DEFAULT_FREEZE_PANES = "A2"
DEFAULT_HEADER_FILL = "1F4E78"
DEFAULT_HEADER_FONT_COLOR = "FFFFFF"
DEFAULT_MAX_COLUMN_WIDTH = 40


# =================================================
# 1. Styling Helpers
# =================================================


def _style_header_row(worksheet, header_fill: str, header_font_color: str) -> None:
    fill = PatternFill(fill_type="solid", fgColor=header_fill)
    font = Font(color=header_font_color, bold=True)
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font


def _autosize_worksheet(worksheet, max_width: int) -> None:
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            cell_text = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_text))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, max_width)


# =================================================
# 2. Workbook Export
# =================================================


def export_stage_workbook(
    workbook_path: Path,
    sheets: dict[str, pd.DataFrame],
    freeze_panes: str = DEFAULT_FREEZE_PANES,
    header_fill: str = DEFAULT_HEADER_FILL,
    header_font_color: str = DEFAULT_HEADER_FONT_COLOR,
    max_column_width: int = DEFAULT_MAX_COLUMN_WIDTH,
) -> Path:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    logger.info("Excel export 시작: %s", workbook_path)

    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        for sheet_name, dataframe in tqdm(sheets.items(), desc="Writing Excel sheets"):
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

    workbook = load_workbook(workbook_path)
    for worksheet in tqdm(workbook.worksheets, desc="Styling Excel sheets"):
        worksheet.freeze_panes = freeze_panes
        if worksheet.max_row >= 1:
            _style_header_row(worksheet, header_fill, header_font_color)
        _autosize_worksheet(worksheet, max_column_width)
    workbook.save(workbook_path)
    logger.info("Excel export 완료: %s", workbook_path)
    return workbook_path
