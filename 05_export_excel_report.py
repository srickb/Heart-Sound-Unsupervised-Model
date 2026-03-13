"""
Assumptions:
- Latest artifacts live under outputs/test_dataset_260312_preprocess_v2.
- Training, clustering, and interpretation outputs already exist.

Expected inputs:
- outputs/{RUN_NAME}/preprocess/*
- outputs/{RUN_NAME}/training/*
- outputs/{RUN_NAME}/clustering/*
- outputs/{RUN_NAME}/interpretation/*

Files I will create:
- outputs/{RUN_NAME}/excel_exports/unsupervised_pipeline_report.xlsx
"""

from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PATHS = {
    "repo_root": Path(__file__).resolve().parent,
}

RUN_NAME = "test_dataset_260312_preprocess_v2"

DATA = {
    "sheet_name_overview": "Overview",
    "sheet_name_history": "Training_History",
    "sheet_name_recon": "Recon_Error",
    "sheet_name_embedding": "Embedding_View",
    "sheet_name_cycle": "Cycle_Level_View",
    "sheet_name_feature_names": "Feature_Names",
    "sheet_name_cluster_summary": "Cluster_Summary",
    "sheet_name_cluster_groups": "Cluster_Groups",
    "sheet_name_cluster_stats": "Cluster_Feature_Stats",
    "sheet_name_repr": "Representative_Samples",
    "sheet_name_noise": "Noise_Analysis",
    "sheet_name_recording": "Recording_Summary",
}

EXCEL = {
    "freeze_header_row": "A2",
    "header_fill": "1F4E78",
    "header_font_color": "FFFFFF",
    "section_fill": "D9EAF7",
    "output_filename": "unsupervised_pipeline_report.xlsx",
}


def build_output_paths() -> dict[str, Path]:
    output_root = PATHS["repo_root"] / "outputs" / RUN_NAME
    export_root = output_root / "excel_exports"
    export_root.mkdir(parents=True, exist_ok=True)
    return {
        "output_root": output_root,
        "preprocess_root": output_root / "preprocess",
        "training_root": output_root / "training",
        "clustering_root": output_root / "clustering",
        "interpretation_root": output_root / "interpretation",
        "export_root": export_root,
        "workbook_path": export_root / EXCEL["output_filename"],
    }


def load_json(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def make_overview_dataframe(
    preprocess_summary: dict,
    training_summary: dict,
    clustering_summary: dict,
) -> pd.DataFrame:
    rows: list[dict[str, object]] = [
        {"section": "Run", "metric": "run_name", "value": RUN_NAME},
        {"section": "Preprocess", "metric": "num_input_files", "value": preprocess_summary["num_input_files"]},
        {"section": "Preprocess", "metric": "num_candidate_cycles", "value": preprocess_summary["num_candidate_cycles"]},
        {"section": "Preprocess", "metric": "num_valid_cycles", "value": preprocess_summary["num_valid_cycles"]},
        {"section": "Preprocess", "metric": "num_excluded_cycles", "value": preprocess_summary["num_excluded_cycles"]},
        {"section": "Preprocess", "metric": "feature_rows", "value": preprocess_summary["feature_shape"][0]},
        {"section": "Preprocess", "metric": "feature_dim", "value": preprocess_summary["feature_shape"][1]},
        {"section": "Training", "metric": "input_dimension", "value": training_summary["input_dimension"]},
        {"section": "Training", "metric": "latent_dimension", "value": training_summary["latent_dimension"]},
        {"section": "Training", "metric": "num_training_samples", "value": training_summary["num_training_samples"]},
        {"section": "Training", "metric": "num_validation_samples", "value": training_summary["num_validation_samples"]},
        {"section": "Training", "metric": "best_validation_loss", "value": training_summary["best_validation_loss"]},
        {"section": "Training", "metric": "final_training_loss", "value": training_summary["final_training_loss"]},
        {"section": "Training", "metric": "final_validation_loss", "value": training_summary["final_validation_loss"]},
        {"section": "Clustering", "metric": "total_valid_samples", "value": clustering_summary["total_valid_samples"]},
        {
            "section": "Clustering",
            "metric": "number_of_clusters_excluding_noise",
            "value": clustering_summary["number_of_clusters_excluding_noise"],
        },
        {"section": "Clustering", "metric": "number_of_noise_samples", "value": clustering_summary["number_of_noise_samples"]},
        {"section": "Clustering", "metric": "noise_ratio", "value": clustering_summary["noise_ratio"]},
        {
            "section": "Clustering",
            "metric": "pca_explained_variance_ratio_pc1",
            "value": clustering_summary["pca_explained_variance_ratio"][0],
        },
        {
            "section": "Clustering",
            "metric": "pca_explained_variance_ratio_pc2",
            "value": clustering_summary["pca_explained_variance_ratio"][1],
        },
    ]
    return pd.DataFrame(rows)


def make_recording_summary_dataframe(preprocess_summary: dict) -> pd.DataFrame:
    recording_rows: list[dict[str, object]] = []
    for record in preprocess_summary["recordings"]:
        flat_row = {key: value for key, value in record.items() if key != "exclusion_reasons"}
        flat_row["exclusion_reasons_json"] = json.dumps(record.get("exclusion_reasons", {}), ensure_ascii=False)
        recording_rows.append(flat_row)
    return pd.DataFrame(recording_rows)


def make_embedding_view(clustering_root: Path) -> pd.DataFrame:
    embedding_meta = pd.read_csv(clustering_root / "embedding_metadata.csv")
    cluster_assignments = pd.read_csv(clustering_root / "cluster_assignments.csv")
    embeddings = np.load(clustering_root / "embeddings.npy")

    embedding_columns = [f"embedding_dim_{index}" for index in range(embeddings.shape[1])]
    embedding_df = pd.DataFrame(embeddings, columns=embedding_columns)
    merged = pd.concat([embedding_meta.reset_index(drop=True), cluster_assignments.drop(columns=["sample_id", "subject_id", "recording_id", "feature_row_index", "waveform_row_index"], errors="ignore").reset_index(drop=True), embedding_df], axis=1)
    return merged


def make_cycle_level_view(output_paths: dict[str, Path]) -> pd.DataFrame:
    metadata = pd.read_csv(output_paths["preprocess_root"] / "cycle_metadata.csv")
    valid_metadata = metadata.loc[metadata["valid_flag"]].copy()

    reconstruction = pd.read_csv(output_paths["training_root"] / "reconstruction_error_summary.csv")
    assignments = pd.read_csv(output_paths["clustering_root"] / "cluster_assignments.csv")

    merged = valid_metadata.merge(
        reconstruction,
        on=["sample_id", "subject_id", "recording_id", "feature_row_index", "waveform_row_index"],
        how="left",
    )
    merged = merged.merge(
        assignments[
            [
                "sample_id",
                "cluster_label",
                "membership_probability",
                "outlier_score",
            ]
        ],
        on="sample_id",
        how="left",
    )
    return merged.sort_values(["feature_row_index"]).reset_index(drop=True)


def make_feature_names_view(preprocess_root: Path) -> pd.DataFrame:
    feature_metadata_path = preprocess_root / "feature_metadata.json"
    feature_names_path = preprocess_root / "feature_names.json"

    if feature_metadata_path.exists():
        with feature_metadata_path.open("r", encoding="utf-8") as handle:
            return pd.DataFrame(json.load(handle))

    with feature_names_path.open("r", encoding="utf-8") as handle:
        feature_names = json.load(handle)
    return pd.DataFrame(
        {
            "feature_index": np.arange(len(feature_names), dtype=int),
            "feature_name": feature_names,
        }
    )


def write_dataframe_sheet(writer: pd.ExcelWriter, dataframe: pd.DataFrame, sheet_name: str) -> None:
    dataframe.to_excel(writer, sheet_name=sheet_name, index=False)


def autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)


def style_header_row(worksheet) -> None:
    fill = PatternFill(fill_type="solid", fgColor=EXCEL["header_fill"])
    font = Font(color=EXCEL["header_font_color"], bold=True)
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font


def add_overview_charts(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)

    overview_sheet = workbook[DATA["sheet_name_overview"]]
    history_sheet = workbook[DATA["sheet_name_history"]]
    cluster_sheet = workbook[DATA["sheet_name_cluster_summary"]]

    line_chart = LineChart()
    line_chart.title = "Training vs Validation Loss"
    line_chart.y_axis.title = "Loss"
    line_chart.x_axis.title = "Epoch"
    data_ref = Reference(history_sheet, min_col=2, max_col=3, min_row=1, max_row=history_sheet.max_row)
    cats_ref = Reference(history_sheet, min_col=1, min_row=2, max_row=history_sheet.max_row)
    line_chart.add_data(data_ref, titles_from_data=True)
    line_chart.set_categories(cats_ref)
    line_chart.height = 7
    line_chart.width = 12
    overview_sheet.add_chart(line_chart, "E2")

    bar_chart = BarChart()
    bar_chart.title = "Cluster Sample Counts"
    bar_chart.y_axis.title = "Samples"
    bar_chart.x_axis.title = "Cluster"
    bar_data = Reference(cluster_sheet, min_col=3, max_col=3, min_row=1, max_row=cluster_sheet.max_row)
    bar_cats = Reference(cluster_sheet, min_col=2, min_row=2, max_row=cluster_sheet.max_row)
    bar_chart.add_data(bar_data, titles_from_data=True)
    bar_chart.set_categories(bar_cats)
    bar_chart.height = 7
    bar_chart.width = 12
    overview_sheet.add_chart(bar_chart, "E20")

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = EXCEL["freeze_header_row"]
        style_header_row(worksheet)
        autosize_worksheet(worksheet)

    workbook.save(workbook_path)


def export_excel_report() -> Path:
    output_paths = build_output_paths()

    preprocess_summary = load_json(output_paths["preprocess_root"] / "preprocess_summary.json")
    training_summary = load_json(output_paths["training_root"] / "training_summary.json")
    clustering_summary = load_json(output_paths["clustering_root"] / "clustering_summary.json")

    overview_df = make_overview_dataframe(preprocess_summary, training_summary, clustering_summary)
    recording_summary_df = make_recording_summary_dataframe(preprocess_summary)
    training_history_df = pd.read_csv(output_paths["training_root"] / "training_history.csv")
    recon_df = pd.read_csv(output_paths["training_root"] / "reconstruction_error_summary.csv")
    embedding_view_df = make_embedding_view(output_paths["clustering_root"])
    cycle_level_df = make_cycle_level_view(output_paths)
    feature_names_df = make_feature_names_view(output_paths["preprocess_root"])
    cluster_summary_df = pd.read_csv(output_paths["interpretation_root"] / "cluster_summary.csv")
    cluster_group_df = pd.read_csv(output_paths["interpretation_root"] / "cluster_group_summary.csv")
    cluster_feature_stats_df = pd.read_csv(output_paths["interpretation_root"] / "cluster_feature_stats.csv")
    representative_df = pd.read_csv(output_paths["interpretation_root"] / "representative_samples.csv")
    noise_df = pd.read_csv(output_paths["interpretation_root"] / "noise_analysis.csv")

    workbook_path = output_paths["workbook_path"]
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        write_dataframe_sheet(writer, overview_df, DATA["sheet_name_overview"])
        write_dataframe_sheet(writer, training_history_df, DATA["sheet_name_history"])
        write_dataframe_sheet(writer, recon_df, DATA["sheet_name_recon"])
        write_dataframe_sheet(writer, embedding_view_df, DATA["sheet_name_embedding"])
        write_dataframe_sheet(writer, cycle_level_df, DATA["sheet_name_cycle"])
        write_dataframe_sheet(writer, feature_names_df, DATA["sheet_name_feature_names"])
        write_dataframe_sheet(writer, cluster_summary_df, DATA["sheet_name_cluster_summary"])
        write_dataframe_sheet(writer, cluster_group_df, DATA["sheet_name_cluster_groups"])
        write_dataframe_sheet(writer, cluster_feature_stats_df, DATA["sheet_name_cluster_stats"])
        write_dataframe_sheet(writer, representative_df, DATA["sheet_name_repr"])
        write_dataframe_sheet(writer, noise_df, DATA["sheet_name_noise"])
        write_dataframe_sheet(writer, recording_summary_df, DATA["sheet_name_recording"])

    add_overview_charts(workbook_path)
    return workbook_path


def main() -> None:
    workbook_path = export_excel_report()
    print(f"Saved Excel report to: {workbook_path}")


if __name__ == "__main__":
    main()
