"""
Standalone preprocessing script for unsupervised heart sound cycle analysis.

Expected input schema per Excel file:
- Time_Index
- Amplitude
- S1-Start_RS_Score
- S1-End_RS_Score
- S2-Start_RS_Score
- S2-End_RS_Score

Saved artifacts:
- outputs/{RUN_NAME}/preprocess/cycle_features.npy
- outputs/{RUN_NAME}/preprocess/cycle_waveforms.npy
- outputs/{RUN_NAME}/preprocess/cycle_metadata.csv
- outputs/{RUN_NAME}/preprocess/feature_names.json
- outputs/{RUN_NAME}/preprocess/feature_metadata.json
- outputs/{RUN_NAME}/preprocess/preprocess_summary.json
- outputs/{RUN_NAME}/preprocess/qc/*
"""

from __future__ import annotations

import json
import logging
import random
from collections import Counter
from pathlib import Path
from typing import Any

import matplotlib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.signal import butter, sosfiltfilt

from excel_export_utils import export_stage_workbook

matplotlib.use("Agg")
import matplotlib.pyplot as plt


# ============================================================================
# Editable configuration
# ============================================================================
class PreprocessConfig:
    PROJECT_ROOT = Path(__file__).resolve().parent
    DATA_ROOT = PROJECT_ROOT / "Data" / "Test_Dataset_260312"
    OUTPUT_ROOT = PROJECT_ROOT / "outputs"
    RUN_NAME = "test_dataset_260312_preprocess_v2"

    FILE_GLOB = "*.xlsx"
    SAMPLING_RATE = 4000
    FIXED_LENGTH = 4000
    MIN_CYCLE_SECONDS = 0.25
    MAX_CYCLE_SECONDS = 1.50

    EXPECTED_COLUMNS = [
        "Time_Index",
        "Amplitude",
        "S1-Start_RS_Score",
        "S1-End_RS_Score",
        "S2-Start_RS_Score",
        "S2-End_RS_Score",
    ]
    RS_SCORE_COLUMNS = [
        "S1-Start_RS_Score",
        "S1-End_RS_Score",
        "S2-Start_RS_Score",
        "S2-End_RS_Score",
    ]

    ENABLE_BANDPASS = False
    BANDPASS_LOW_HZ = 20.0
    BANDPASS_HIGH_HZ = 800.0
    BANDPASS_ORDER = 4

    SUBTRACT_CYCLE_MEAN_BEFORE_NORMALIZATION = True
    NORMALIZE_WAVEFORM_BY = "max_abs"

    RS_LOCAL_WINDOW_RADIUS = 20
    RS_INCLUDE_LOCAL_WINDOW_STATS = True

    FEATURE_GROUP_MODE = "all"
    INCLUDE_RATIO_FEATURES_IN_ABLATION = True
    INCLUDE_QC_FEATURES_IN_ABLATION = True

    ENABLE_RS_FEATURES = True
    ENABLE_TIME_FEATURES = True
    ENABLE_AMPLITUDE_FEATURES = True
    ENABLE_RAW_AMPLITUDE_FEATURES = True
    ENABLE_NORMALIZED_AMPLITUDE_FEATURES = True
    ENABLE_RATIO_FEATURES = True
    ENABLE_QC_FEATURES = True

    HARD_EXCLUDE_MISSING_EVENTS = True
    HARD_EXCLUDE_UNORDERABLE_EVENTS = True
    HARD_EXCLUDE_EMPTY_SEGMENTS = True
    HARD_EXCLUDE_OUT_OF_BOUNDS_CYCLES = False

    QC_NUM_EXAMPLE_WAVEFORMS = 8
    QC_WAVEFORM_ALPHA = 0.7

    EXCEL_EXPORT_ENABLED = True
    EXCEL_FILENAME = "preprocess_data_export.xlsx"
    EXCEL_FREEZE_PANES = "A2"
    EXCEL_HEADER_FILL = "1F4E78"
    EXCEL_HEADER_FONT_COLOR = "FFFFFF"
    EXCEL_MAX_COLUMN_WIDTH = 40

    RANDOM_SEED = 42


logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

PATHS = {
    "data_root": PreprocessConfig.DATA_ROOT,
    "output_root": PreprocessConfig.OUTPUT_ROOT,
}

RUN_NAME = PreprocessConfig.RUN_NAME

DATA = {
    "file_glob": PreprocessConfig.FILE_GLOB,
    "sampling_rate": PreprocessConfig.SAMPLING_RATE,
    "expected_columns": PreprocessConfig.EXPECTED_COLUMNS,
}

PREPROCESS = {
    "enable_bandpass": PreprocessConfig.ENABLE_BANDPASS,
    "bandpass_low_hz": PreprocessConfig.BANDPASS_LOW_HZ,
    "bandpass_high_hz": PreprocessConfig.BANDPASS_HIGH_HZ,
    "bandpass_order": PreprocessConfig.BANDPASS_ORDER,
    "fixed_length": PreprocessConfig.FIXED_LENGTH,
    "min_cycle_seconds": PreprocessConfig.MIN_CYCLE_SECONDS,
    "max_cycle_seconds": PreprocessConfig.MAX_CYCLE_SECONDS,
}

FEATURE = {
    "normalize_waveform_by": PreprocessConfig.NORMALIZE_WAVEFORM_BY,
    "subtract_cycle_mean_before_normalization": PreprocessConfig.SUBTRACT_CYCLE_MEAN_BEFORE_NORMALIZATION,
    "rs_score_columns": PreprocessConfig.RS_SCORE_COLUMNS,
    "rs_local_window_radius": PreprocessConfig.RS_LOCAL_WINDOW_RADIUS,
    "rs_include_local_window_stats": PreprocessConfig.RS_INCLUDE_LOCAL_WINDOW_STATS,
    "feature_group_mode": PreprocessConfig.FEATURE_GROUP_MODE,
    "include_ratio_features_in_ablation": PreprocessConfig.INCLUDE_RATIO_FEATURES_IN_ABLATION,
    "include_qc_features_in_ablation": PreprocessConfig.INCLUDE_QC_FEATURES_IN_ABLATION,
    "enable_rs_features": PreprocessConfig.ENABLE_RS_FEATURES,
    "enable_time_features": PreprocessConfig.ENABLE_TIME_FEATURES,
    "enable_amplitude_features": PreprocessConfig.ENABLE_AMPLITUDE_FEATURES,
    "enable_raw_amplitude_features": PreprocessConfig.ENABLE_RAW_AMPLITUDE_FEATURES,
    "enable_normalized_amplitude_features": PreprocessConfig.ENABLE_NORMALIZED_AMPLITUDE_FEATURES,
    "enable_ratio_features": PreprocessConfig.ENABLE_RATIO_FEATURES,
    "enable_qc_features": PreprocessConfig.ENABLE_QC_FEATURES,
}

QC = {
    "hard_exclude_missing_events": PreprocessConfig.HARD_EXCLUDE_MISSING_EVENTS,
    "hard_exclude_unorderable_events": PreprocessConfig.HARD_EXCLUDE_UNORDERABLE_EVENTS,
    "hard_exclude_empty_segments": PreprocessConfig.HARD_EXCLUDE_EMPTY_SEGMENTS,
    "hard_exclude_out_of_bounds_cycles": PreprocessConfig.HARD_EXCLUDE_OUT_OF_BOUNDS_CYCLES,
    "num_example_waveforms": PreprocessConfig.QC_NUM_EXAMPLE_WAVEFORMS,
    "waveform_alpha": PreprocessConfig.QC_WAVEFORM_ALPHA,
}

EXCEL = {
    "export_enabled": PreprocessConfig.EXCEL_EXPORT_ENABLED,
    "filename": PreprocessConfig.EXCEL_FILENAME,
    "freeze_panes": PreprocessConfig.EXCEL_FREEZE_PANES,
    "header_fill": PreprocessConfig.EXCEL_HEADER_FILL,
    "header_font_color": PreprocessConfig.EXCEL_HEADER_FONT_COLOR,
    "max_column_width": PreprocessConfig.EXCEL_MAX_COLUMN_WIDTH,
}

RANDOM_SEED = PreprocessConfig.RANDOM_SEED

EVENT_COLUMN_MAP = {
    "s1_start": "S1-Start_RS_Score",
    "s1_end": "S1-End_RS_Score",
    "s2_start": "S2-Start_RS_Score",
    "s2_end": "S2-End_RS_Score",
}
EVENT_SEQUENCE = ["s1_start", "s1_end", "s2_start", "s2_end"]
EVENT_LABELS = {
    "s1_start": "S1_START",
    "s1_end": "S1_END",
    "s2_start": "S2_START",
    "s2_end": "S2_END",
}
RS_SCORE_NAME_MAP = {
    "S1-Start_RS_Score": "s1_start_score",
    "S1-End_RS_Score": "s1_end_score",
    "S2-Start_RS_Score": "s2_start_score",
    "S2-End_RS_Score": "s2_end_score",
}
FEATURE_PREFIX_TO_GROUP = {
    "rs_": "rs",
    "time_": "time",
    "amp_raw_": "amp_raw",
    "amp_norm_": "amp_norm",
    "ratio_": "ratio",
    "qc_": "qc",
}


# ============================================================================
# Dataset adapter section
# ============================================================================
def parse_recording_identity(file_path: Path) -> dict[str, str]:
    """Parse subject and site identifiers from the recording filename."""
    stem_parts = file_path.stem.split("_")
    if len(stem_parts) < 4:
        raise ValueError(
            f"Unexpected filename format for recording identity: {file_path.name}"
        )

    subject_id = stem_parts[0]
    auscultation_site = stem_parts[1]
    recording_id = file_path.stem
    return {
        "recording_id": recording_id,
        "subject_id": subject_id,
        "auscultation_site": auscultation_site,
    }


def load_recording_table(
    file_path: Path,
    expected_columns: list[str],
) -> pd.DataFrame:
    """Load one workbook and enforce the expected column schema."""
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)

    try:
        header = [str(value) if value is not None else "" for value in next(rows)]
    except StopIteration as exc:
        workbook.close()
        raise ValueError(f"Workbook is empty: {file_path}") from exc

    missing_columns = [column for column in expected_columns if column not in header]
    if missing_columns:
        workbook.close()
        raise ValueError(
            f"Missing required columns in {file_path.name}: {missing_columns}"
        )

    column_indices = [header.index(column) for column in expected_columns]
    data = {column: [] for column in expected_columns}

    for row in rows:
        if row is None:
            continue
        values = [row[index] if index < len(row) else None for index in column_indices]
        if all(value is None for value in values):
            continue
        for column, value in zip(expected_columns, values):
            data[column].append(value)

    workbook.close()

    dataframe = pd.DataFrame(data)
    if dataframe.empty:
        raise ValueError(f"No data rows found in workbook: {file_path}")

    dataframe = dataframe.replace({None: np.nan})
    if dataframe[expected_columns].isnull().any().any():
        null_counts = dataframe[expected_columns].isnull().sum()
        failing = null_counts[null_counts > 0].to_dict()
        raise ValueError(f"Null values detected in {file_path.name}: {failing}")

    dataframe["Time_Index"] = dataframe["Time_Index"].astype(np.int64)
    dataframe["Amplitude"] = dataframe["Amplitude"].astype(np.float32)
    for column in expected_columns[2:]:
        dataframe[column] = dataframe[column].astype(np.float32)

    if len(dataframe) < 2:
        raise ValueError(f"Recording is too short to segment cycles: {file_path.name}")

    time_index = dataframe["Time_Index"].to_numpy(dtype=np.int64)
    time_diffs = np.diff(time_index)
    if np.any(time_diffs <= 0):
        raise ValueError(f"Non-monotonic Time_Index detected in {file_path.name}")
    if np.any(time_diffs != 1):
        raise ValueError(
            f"Time_Index must increment by 1 sample in {file_path.name}, "
            f"but observed diffs {np.unique(time_diffs)}"
        )

    return dataframe


# ============================================================================
# Utility functions
# ============================================================================
def set_random_seed(seed: int) -> None:
    """Set deterministic seeds for Python and NumPy."""
    random.seed(seed)
    np.random.seed(seed)


def ensure_output_directories(output_root: Path, run_name: str) -> dict[str, Path]:
    """Create stable output directories for this run."""
    run_root = output_root / run_name
    preprocess_root = run_root / "preprocess"
    qc_root = preprocess_root / "qc"
    training_root = run_root / "training"
    clustering_root = run_root / "clustering"
    interpretation_root = run_root / "interpretation"

    for path in [
        preprocess_root,
        qc_root,
        training_root,
        clustering_root,
        interpretation_root,
    ]:
        path.mkdir(parents=True, exist_ok=True)

    return {
        "run_root": run_root,
        "preprocess_root": preprocess_root,
        "qc_root": qc_root,
        "training_root": training_root,
        "clustering_root": clustering_root,
        "interpretation_root": interpretation_root,
    }


def maybe_filter_signal(
    waveform: np.ndarray,
    sampling_rate: int,
    preprocess_config: dict[str, Any],
) -> np.ndarray:
    """Apply optional conservative bandpass filtering."""
    waveform = waveform.astype(np.float32, copy=True)
    if not preprocess_config["enable_bandpass"]:
        return waveform

    sos = butter(
        preprocess_config["bandpass_order"],
        [
            preprocess_config["bandpass_low_hz"],
            preprocess_config["bandpass_high_hz"],
        ],
        btype="bandpass",
        output="sos",
        fs=sampling_rate,
    )
    return sosfiltfilt(sos, waveform).astype(np.float32)


def find_positive_runs(score_array: np.ndarray) -> list[tuple[int, int]]:
    """Find contiguous positive-score runs."""
    positive_rows = np.flatnonzero(score_array > 0)
    if positive_rows.size == 0:
        return []

    runs: list[tuple[int, int]] = []
    run_start = int(positive_rows[0])
    previous_row = int(positive_rows[0])
    for row_index in positive_rows[1:]:
        row_index = int(row_index)
        if row_index == previous_row + 1:
            previous_row = row_index
            continue
        runs.append((run_start, previous_row))
        run_start = row_index
        previous_row = row_index
    runs.append((run_start, previous_row))
    return runs


def representative_peak_row(score_array: np.ndarray, start_row: int, end_row: int) -> int:
    """Pick one representative row from a positive RS-score run."""
    run_scores = score_array[start_row : end_row + 1]
    max_score = np.max(run_scores)
    plateau_rows = np.flatnonzero(run_scores == max_score)
    plateau_center = int(plateau_rows[len(plateau_rows) // 2])
    return start_row + plateau_center


def extract_event_candidates(
    time_index: np.ndarray,
    score_array: np.ndarray,
    event_name: str,
) -> list[dict[str, Any]]:
    """Convert RS-score runs into cycle event candidates."""
    candidates: list[dict[str, Any]] = []
    for run_start_row, run_end_row in find_positive_runs(score_array):
        peak_row = representative_peak_row(score_array, run_start_row, run_end_row)
        candidates.append(
            {
                "event_name": event_name,
                "run_start_row": run_start_row,
                "run_end_row": run_end_row,
                "peak_row": peak_row,
                "run_start_sample": int(time_index[run_start_row]),
                "run_end_sample": int(time_index[run_end_row]),
                "sample_index": int(time_index[peak_row]),
                "peak_score": float(score_array[peak_row]),
            }
        )
    return candidates


def events_between(
    events: list[dict[str, Any]],
    left_sample: int,
    right_sample: int,
) -> list[dict[str, Any]]:
    """Return events strictly inside one cycle boundary."""
    return [
        event
        for event in events
        if left_sample < int(event["sample_index"]) < right_sample
    ]


def sanitize_feature_value(value: Any, default: float = 0.0) -> float:
    """Convert feature values to stable finite floats."""
    try:
        numeric_value = float(value)
    except (TypeError, ValueError):
        return float(default)
    if np.isnan(numeric_value) or np.isinf(numeric_value):
        return float(default)
    return numeric_value


def sanitize_feature_dict(features: dict[str, Any], default: float = 0.0) -> dict[str, float]:
    """Convert a feature dictionary into a finite float-only mapping."""
    return {
        key: sanitize_feature_value(value, default=default)
        for key, value in features.items()
    }


def safe_ratio(numerator: float, denominator: float, default: float = 0.0) -> float:
    """Return a finite safe ratio for downstream training."""
    denominator = float(denominator)
    if abs(denominator) < 1e-12:
        return float(default)
    return sanitize_feature_value(float(numerator) / denominator, default=default)


def normalize_cycle_waveform(
    waveform: np.ndarray,
    subtract_mean: bool,
) -> np.ndarray:
    """Normalize one cycle waveform for shape-based features."""
    normalized = waveform.astype(np.float32, copy=True)
    if subtract_mean:
        normalized = normalized - np.mean(normalized)

    max_abs = float(np.max(np.abs(normalized)))
    if max_abs > 0.0:
        normalized = normalized / max_abs
    return normalized.astype(np.float32)


def fixed_length_representation(
    waveform: np.ndarray,
    fixed_length: int,
) -> np.ndarray:
    """Convert one variable-length cycle waveform into a fixed-length array."""
    output = np.zeros(fixed_length, dtype=np.float32)
    copy_length = min(len(waveform), fixed_length)
    output[:copy_length] = waveform[:copy_length]
    return output


def safe_skewness(values: np.ndarray) -> float:
    """Compute a finite skewness-like value without introducing NaN."""
    values = np.asarray(values, dtype=np.float64)
    if values.size < 3:
        return 0.0
    centered = values - np.mean(values)
    std = float(np.std(centered))
    if std == 0.0:
        return 0.0
    third_moment = float(np.mean(np.power(centered / std, 3)))
    return sanitize_feature_value(third_moment)


def safe_kurtosis(values: np.ndarray) -> float:
    """Compute Fisher-style excess kurtosis with finite fallback."""
    values = np.asarray(values, dtype=np.float64)
    if values.size < 4:
        return 0.0
    centered = values - np.mean(values)
    std = float(np.std(centered))
    if std == 0.0:
        return 0.0
    fourth_moment = float(np.mean(np.power(centered / std, 4)) - 3.0)
    return sanitize_feature_value(fourth_moment)


def waveform_statistics(
    waveform: np.ndarray,
    prefix: str,
) -> dict[str, float]:
    """Compute interpretable statistics from one waveform segment."""
    waveform = np.asarray(waveform, dtype=np.float32)
    if waveform.size == 0:
        return {
            f"{prefix}_mean": 0.0,
            f"{prefix}_std": 0.0,
            f"{prefix}_max": 0.0,
            f"{prefix}_min": 0.0,
            f"{prefix}_max_abs": 0.0,
            f"{prefix}_peak_to_peak": 0.0,
            f"{prefix}_rms": 0.0,
            f"{prefix}_energy": 0.0,
            f"{prefix}_abs_area": 0.0,
            f"{prefix}_skewness": 0.0,
            f"{prefix}_kurtosis": 0.0,
        }

    stats: dict[str, float] = {
        f"{prefix}_mean": float(np.mean(waveform)),
        f"{prefix}_std": float(np.std(waveform)),
        f"{prefix}_max": float(np.max(waveform)),
        f"{prefix}_min": float(np.min(waveform)),
        f"{prefix}_max_abs": float(np.max(np.abs(waveform))),
        f"{prefix}_peak_to_peak": float(np.ptp(waveform)),
        f"{prefix}_rms": float(np.sqrt(np.mean(np.square(waveform)))),
        f"{prefix}_energy": float(np.sum(np.square(waveform))),
        f"{prefix}_abs_area": float(np.sum(np.abs(waveform))),
        f"{prefix}_skewness": safe_skewness(waveform),
        f"{prefix}_kurtosis": safe_kurtosis(waveform),
    }
    return sanitize_feature_dict(stats)


def feature_group_from_name(feature_name: str) -> str:
    """Infer the semantic feature group from its prefix."""
    for prefix, group_name in FEATURE_PREFIX_TO_GROUP.items():
        if feature_name.startswith(prefix):
            return group_name
    return "other"


def summarize_array(values: np.ndarray) -> dict[str, float] | dict[str, None]:
    """Return min/median/mean/max summary for a 1D numeric array."""
    if values.size == 0:
        return {"min": None, "median": None, "mean": None, "max": None}
    return {
        "min": float(np.min(values)),
        "median": float(np.median(values)),
        "mean": float(np.mean(values)),
        "max": float(np.max(values)),
    }


def summarize_boolean_array(values: np.ndarray) -> dict[str, float | int]:
    """Summarize boolean flags as counts and ratios."""
    values = np.asarray(values, dtype=bool)
    if values.size == 0:
        return {"count_true": 0, "count_false": 0, "ratio_true": 0.0}
    count_true = int(values.sum())
    return {
        "count_true": count_true,
        "count_false": int(values.size - count_true),
        "ratio_true": float(count_true / values.size),
    }


def select_feature_prefixes(feature_config: dict[str, Any]) -> set[str]:
    """Resolve the selected feature prefixes from ablation mode and toggles."""
    mode = str(feature_config["feature_group_mode"]).lower()
    mode_to_prefixes = {
        "all": {"rs_", "time_", "amp_raw_", "amp_norm_"},
        "rs_only": {"rs_"},
        "time_only": {"time_"},
        "amplitude_only": {"amp_raw_", "amp_norm_"},
        "rs_time": {"rs_", "time_"},
        "rs_time_amplitude": {"rs_", "time_", "amp_raw_", "amp_norm_"},
    }
    if mode not in mode_to_prefixes:
        raise ValueError(f"Unsupported feature group mode: {feature_config['feature_group_mode']}")

    prefixes = set(mode_to_prefixes[mode])

    if not feature_config["enable_rs_features"]:
        prefixes.discard("rs_")
    if not feature_config["enable_time_features"]:
        prefixes.discard("time_")
    if not feature_config["enable_amplitude_features"]:
        prefixes.discard("amp_raw_")
        prefixes.discard("amp_norm_")
    else:
        if not feature_config["enable_raw_amplitude_features"]:
            prefixes.discard("amp_raw_")
        if not feature_config["enable_normalized_amplitude_features"]:
            prefixes.discard("amp_norm_")

    if feature_config["enable_ratio_features"]:
        if mode == "all" or feature_config["include_ratio_features_in_ablation"]:
            prefixes.add("ratio_")
    if feature_config["enable_qc_features"]:
        if mode == "all" or feature_config["include_qc_features_in_ablation"]:
            prefixes.add("qc_")
    return prefixes


def select_enabled_features(
    all_features: dict[str, float],
    feature_config: dict[str, Any],
) -> dict[str, float]:
    """Keep only the selected feature groups while preserving key order."""
    selected_prefixes = select_feature_prefixes(feature_config)
    filtered = {
        name: value
        for name, value in all_features.items()
        if any(name.startswith(prefix) for prefix in selected_prefixes)
    }
    return sanitize_feature_dict(filtered)


def build_feature_metadata(feature_names: list[str]) -> list[dict[str, Any]]:
    """Create a compact metadata table for feature names and groups."""
    rows: list[dict[str, Any]] = []
    for feature_index, feature_name in enumerate(feature_names):
        rows.append(
            {
                "feature_index": int(feature_index),
                "feature_name": feature_name,
                "feature_group": feature_group_from_name(feature_name),
                "feature_prefix": feature_name.split("_", 1)[0] if "_" in feature_name else "other",
            }
        )
    return rows


def best_valid_event_combination(
    cycle_start_sample: int,
    next_s1_start_sample: int,
    s1_end_candidates: list[dict[str, Any]],
    s2_start_candidates: list[dict[str, Any]],
    s2_end_candidates: list[dict[str, Any]],
) -> tuple[dict[str, Any] | None, str, dict[str, Any]]:
    """
    Resolve cycle events while allowing multiple candidates.

    Missing events remain hard failures by default, but multiple candidates are
    handled by choosing the highest-confidence ordered combination.
    """
    diagnostics = {
        "num_s1_end_candidates": len(s1_end_candidates),
        "num_s2_start_candidates": len(s2_start_candidates),
        "num_s2_end_candidates": len(s2_end_candidates),
        "selected_from_multiple_candidates_flag": int(
            len(s1_end_candidates) > 1
            or len(s2_start_candidates) > 1
            or len(s2_end_candidates) > 1
        ),
    }

    if len(s1_end_candidates) == 0:
        return None, "missing_s1_end", diagnostics
    if len(s2_start_candidates) == 0:
        return None, "missing_s2_start", diagnostics
    if len(s2_end_candidates) == 0:
        return None, "missing_s2_end", diagnostics

    valid_combinations: list[tuple[tuple[float, float, float, float], dict[str, Any]]] = []
    cycle_midpoint = 0.5 * (cycle_start_sample + next_s1_start_sample)

    for s1_end_event in s1_end_candidates:
        s1_end_sample = int(s1_end_event["sample_index"])
        if not cycle_start_sample < s1_end_sample < next_s1_start_sample:
            continue
        for s2_start_event in s2_start_candidates:
            s2_start_sample = int(s2_start_event["sample_index"])
            if not s1_end_sample < s2_start_sample < next_s1_start_sample:
                continue
            for s2_end_event in s2_end_candidates:
                s2_end_sample = int(s2_end_event["sample_index"])
                if not s2_start_sample < s2_end_sample < next_s1_start_sample:
                    continue

                total_peak_score = (
                    float(s1_end_event["peak_score"])
                    + float(s2_start_event["peak_score"])
                    + float(s2_end_event["peak_score"])
                )
                total_run_width = (
                    int(s1_end_event["run_end_row"]) - int(s1_end_event["run_start_row"])
                    + int(s2_start_event["run_end_row"]) - int(s2_start_event["run_start_row"])
                    + int(s2_end_event["run_end_row"]) - int(s2_end_event["run_start_row"])
                )
                midpoint_distance = (
                    abs(s1_end_sample - cycle_midpoint)
                    + abs(s2_start_sample - cycle_midpoint)
                    + abs(s2_end_sample - cycle_midpoint)
                )
                sort_key = (
                    total_peak_score,
                    -float(total_run_width),
                    -float(midpoint_distance),
                    -float(s2_end_sample),
                )
                valid_combinations.append(
                    (
                        sort_key,
                        {
                            "s1_end": s1_end_event,
                            "s2_start": s2_start_event,
                            "s2_end": s2_end_event,
                        },
                    )
                )

    if not valid_combinations:
        return None, "invalid_event_order", diagnostics

    valid_combinations.sort(key=lambda item: item[0], reverse=True)
    selected_events = valid_combinations[0][1]
    diagnostics["event_selection_strategy"] = (
        "unique_candidates"
        if diagnostics["selected_from_multiple_candidates_flag"] == 0
        else "best_ordered_peak_score"
    )
    return selected_events, "", diagnostics


def extract_rs_landmark_features(
    recording_table: pd.DataFrame,
    event_rows: dict[str, int],
    rs_columns: list[str],
    local_window_radius: int,
    include_local_window_stats: bool,
) -> dict[str, float]:
    """Build the required RS landmark feature set for one cycle."""
    features: dict[str, float] = {}
    for anchor_event in EVENT_SEQUENCE:
        anchor_row = int(event_rows[anchor_event])
        anchor_name = anchor_event
        for score_column in rs_columns:
            score_name = RS_SCORE_NAME_MAP[score_column]
            value = recording_table.iloc[anchor_row][score_column]
            features[f"rs_at_{anchor_name}__{score_name}"] = float(value)

            if include_local_window_stats:
                left_row = max(0, anchor_row - local_window_radius)
                right_row = min(len(recording_table), anchor_row + local_window_radius + 1)
                window = recording_table.iloc[left_row:right_row][score_column].to_numpy(dtype=np.float32)
                features[f"rs_window_at_{anchor_name}__{score_name}_local_max"] = float(np.max(window))
                features[f"rs_window_at_{anchor_name}__{score_name}_local_mean"] = float(np.mean(window))
                features[f"rs_window_at_{anchor_name}__{score_name}_local_std"] = float(np.std(window))

    return sanitize_feature_dict(features)


def build_time_features(
    cycle_start_sample: int,
    next_s1_start_sample: int,
    s1_end_sample: int,
    s2_start_sample: int,
    s2_end_sample: int,
    sampling_rate: int,
) -> dict[str, float]:
    """Build interpretable cycle timing features with explicit prefixes."""
    cycle_duration_sec = (next_s1_start_sample - cycle_start_sample) / sampling_rate
    s1_duration_sec = (s1_end_sample - cycle_start_sample) / sampling_rate
    systole_duration_sec = (s2_start_sample - s1_end_sample) / sampling_rate
    s2_duration_sec = (s2_end_sample - s2_start_sample) / sampling_rate
    diastole_duration_sec = (next_s1_start_sample - s2_end_sample) / sampling_rate
    s1_to_s2_interval_sec = (s2_start_sample - cycle_start_sample) / sampling_rate
    s2_to_next_s1_interval_sec = (next_s1_start_sample - s2_start_sample) / sampling_rate

    s1_center = 0.5 * (cycle_start_sample + s1_end_sample)
    s2_center = 0.5 * (s2_start_sample + s2_end_sample)

    features = {
        "time_cycle_duration_sec": cycle_duration_sec,
        "time_s1_duration_sec": s1_duration_sec,
        "time_systole_duration_sec": systole_duration_sec,
        "time_s2_duration_sec": s2_duration_sec,
        "time_diastole_duration_sec": diastole_duration_sec,
        "time_s1_to_s2_interval_sec": s1_to_s2_interval_sec,
        "time_s2_to_next_s1_interval_sec": s2_to_next_s1_interval_sec,
        "time_s1_ratio": safe_ratio(s1_duration_sec, cycle_duration_sec),
        "time_systole_ratio": safe_ratio(systole_duration_sec, cycle_duration_sec),
        "time_s2_ratio": safe_ratio(s2_duration_sec, cycle_duration_sec),
        "time_diastole_ratio": safe_ratio(diastole_duration_sec, cycle_duration_sec),
        "time_s1_to_s2_ratio": safe_ratio(s1_to_s2_interval_sec, cycle_duration_sec),
        "time_s2_to_next_s1_ratio": safe_ratio(s2_to_next_s1_interval_sec, cycle_duration_sec),
        "time_s1_center_to_s2_center_sec": (s2_center - s1_center) / sampling_rate,
        "time_s1_start_to_s2_start_sec": (s2_start_sample - cycle_start_sample) / sampling_rate,
        "time_s1_end_to_s2_end_sec": (s2_end_sample - s1_end_sample) / sampling_rate,
    }
    return sanitize_feature_dict(features)


def build_amplitude_features(
    segment_map: dict[str, np.ndarray],
    prefix: str,
) -> dict[str, float]:
    """Build amplitude statistics for a family of waveform segments."""
    features: dict[str, float] = {}
    for segment_name, segment_waveform in segment_map.items():
        features.update(waveform_statistics(segment_waveform, prefix=f"{prefix}_{segment_name}"))
    return sanitize_feature_dict(features)


def build_ratio_features(all_features: dict[str, float]) -> dict[str, float]:
    """Build comparative ratios across key time and amplitude measurements."""
    features = {
        "ratio_amp_raw_s1_max_abs_over_amp_raw_s2_max_abs": safe_ratio(
            all_features.get("amp_raw_s1_max_abs", 0.0),
            all_features.get("amp_raw_s2_max_abs", 0.0),
        ),
        "ratio_amp_raw_systole_energy_over_amp_raw_diastole_energy": safe_ratio(
            all_features.get("amp_raw_systole_energy", 0.0),
            all_features.get("amp_raw_diastole_energy", 0.0),
        ),
        "ratio_amp_norm_s1_rms_over_amp_norm_s2_rms": safe_ratio(
            all_features.get("amp_norm_s1_rms", 0.0),
            all_features.get("amp_norm_s2_rms", 0.0),
        ),
        "ratio_time_s1_duration_over_time_s2_duration": safe_ratio(
            all_features.get("time_s1_duration_sec", 0.0),
            all_features.get("time_s2_duration_sec", 0.0),
        ),
        "ratio_time_systole_duration_over_time_diastole_duration": safe_ratio(
            all_features.get("time_systole_duration_sec", 0.0),
            all_features.get("time_diastole_duration_sec", 0.0),
        ),
        "ratio_amp_raw_cycle_energy_over_amp_norm_cycle_energy": safe_ratio(
            all_features.get("amp_raw_cycle_energy", 0.0),
            all_features.get("amp_norm_cycle_energy", 0.0),
        ),
    }
    return sanitize_feature_dict(features)


def build_qc_features(
    cycle_length_samples: int,
    sampling_rate: int,
    cycle_length_in_bounds_flag: int,
    event_ordering_valid_flag: int,
    diagnostics: dict[str, Any],
) -> dict[str, float]:
    """Build featureized QC signals instead of discarding every noisy cycle."""
    min_cycle_samples = int(PREPROCESS["min_cycle_seconds"] * sampling_rate)
    max_cycle_samples = int(PREPROCESS["max_cycle_seconds"] * sampling_rate)
    cycle_duration_sec = cycle_length_samples / sampling_rate

    features = {
        "qc_num_s1_end_candidates": float(diagnostics["num_s1_end_candidates"]),
        "qc_num_s2_start_candidates": float(diagnostics["num_s2_start_candidates"]),
        "qc_num_s2_end_candidates": float(diagnostics["num_s2_end_candidates"]),
        "qc_event_ordering_valid_flag": float(event_ordering_valid_flag),
        "qc_cycle_length_in_bounds_flag": float(cycle_length_in_bounds_flag),
        "qc_selected_from_multiple_candidates_flag": float(
            diagnostics["selected_from_multiple_candidates_flag"]
        ),
        "qc_cycle_duration_zless_bounds_margin_sec": min(
            cycle_duration_sec - PREPROCESS["min_cycle_seconds"],
            PREPROCESS["max_cycle_seconds"] - cycle_duration_sec,
        ),
        "qc_cycle_length_below_min_samples": float(max(0, min_cycle_samples - cycle_length_samples)),
        "qc_cycle_length_above_max_samples": float(max(0, cycle_length_samples - max_cycle_samples)),
    }
    return sanitize_feature_dict(features)


def build_full_feature_row(
    recording_table: pd.DataFrame,
    event_rows: dict[str, int],
    raw_segment_map: dict[str, np.ndarray],
    normalized_segment_map: dict[str, np.ndarray],
    event_samples: dict[str, int],
    next_s1_start_sample: int,
    sampling_rate: int,
    cycle_length_in_bounds_flag: int,
    diagnostics: dict[str, Any],
) -> dict[str, float]:
    """Build the full feature row before ablation-based selection."""
    cycle_start_sample = event_samples["s1_start"]
    s1_end_sample = event_samples["s1_end"]
    s2_start_sample = event_samples["s2_start"]
    s2_end_sample = event_samples["s2_end"]
    cycle_length_samples = next_s1_start_sample - cycle_start_sample

    features: dict[str, float] = {}
    features.update(
        extract_rs_landmark_features(
            recording_table=recording_table,
            event_rows=event_rows,
            rs_columns=FEATURE["rs_score_columns"],
            local_window_radius=FEATURE["rs_local_window_radius"],
            include_local_window_stats=FEATURE["rs_include_local_window_stats"],
        )
    )
    features.update(
        build_time_features(
            cycle_start_sample=cycle_start_sample,
            next_s1_start_sample=next_s1_start_sample,
            s1_end_sample=s1_end_sample,
            s2_start_sample=s2_start_sample,
            s2_end_sample=s2_end_sample,
            sampling_rate=sampling_rate,
        )
    )
    features.update(build_amplitude_features(raw_segment_map, prefix="amp_raw"))
    features.update(build_amplitude_features(normalized_segment_map, prefix="amp_norm"))
    features.update(build_ratio_features(features))
    features.update(
        build_qc_features(
            cycle_length_samples=cycle_length_samples,
            sampling_rate=sampling_rate,
            cycle_length_in_bounds_flag=cycle_length_in_bounds_flag,
            event_ordering_valid_flag=1,
            diagnostics=diagnostics,
        )
    )
    return sanitize_feature_dict(features)


def preferred_duration_column(metadata_frame: pd.DataFrame) -> str | None:
    """Find the duration column used for QC plots."""
    for column in ["time_cycle_duration_sec", "cycle_duration_sec"]:
        if column in metadata_frame.columns:
            return column
    return None


def save_qc_plots(
    metadata_frame: pd.DataFrame,
    waveform_matrix: np.ndarray,
    qc_root: Path,
    fixed_length: int,
    random_seed: int,
) -> None:
    """Generate simple QC figures for the preprocessing output."""
    valid_metadata = metadata_frame[metadata_frame["valid_flag"] == True].copy()
    excluded_metadata = metadata_frame[metadata_frame["valid_flag"] == False].copy()
    duration_column = preferred_duration_column(metadata_frame)

    plt.figure(figsize=(8, 4.5))
    if not valid_metadata.empty and duration_column is not None:
        plt.hist(valid_metadata[duration_column], bins=30, color="#4C78A8")
    plt.xlabel("Cycle duration (sec)")
    plt.ylabel("Count")
    plt.title("Valid Cycle Duration Histogram")
    plt.tight_layout()
    plt.savefig(qc_root / "cycle_duration_histogram.png", dpi=150)
    plt.close()

    plt.figure(figsize=(6, 4))
    plt.bar(
        ["valid", "excluded"],
        [len(valid_metadata), len(excluded_metadata)],
        color=["#59A14F", "#E15759"],
    )
    plt.ylabel("Count")
    plt.title("Valid vs Excluded Cycles")
    plt.tight_layout()
    plt.savefig(qc_root / "valid_vs_excluded_cycles.png", dpi=150)
    plt.close()

    plt.figure(figsize=(10, 5))
    if waveform_matrix.size > 0:
        rng = np.random.default_rng(random_seed)
        num_examples = min(QC["num_example_waveforms"], waveform_matrix.shape[0])
        chosen_rows = np.sort(
            rng.choice(waveform_matrix.shape[0], size=num_examples, replace=False)
        )
        x_axis = np.arange(fixed_length)
        for row_index in chosen_rows:
            plt.plot(
                x_axis,
                waveform_matrix[row_index],
                alpha=QC["waveform_alpha"],
                linewidth=1.0,
            )
    plt.xlabel("Sample")
    plt.ylabel("Normalized amplitude")
    plt.title("Example Normalized Cycle Waveforms")
    plt.tight_layout()
    plt.savefig(qc_root / "example_normalized_cycle_waveforms.png", dpi=150)
    plt.close()


# ============================================================================
# Main preprocessing pipeline
# ============================================================================
def process_recording(
    file_path: Path,
    sampling_rate: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    """Process one recording into cycle-level metadata, features, and waveforms."""
    identity = parse_recording_identity(file_path)
    recording_table = load_recording_table(file_path, DATA["expected_columns"])
    filtered_amplitude = maybe_filter_signal(
        recording_table["Amplitude"].to_numpy(dtype=np.float32),
        sampling_rate=sampling_rate,
        preprocess_config=PREPROCESS,
    )
    time_index = recording_table["Time_Index"].to_numpy(dtype=np.int64)

    event_candidates = {
        event_key: extract_event_candidates(
            time_index,
            recording_table[column_name].to_numpy(dtype=np.float32),
            event_name=EVENT_LABELS[event_key],
        )
        for event_key, column_name in EVENT_COLUMN_MAP.items()
    }

    if len(event_candidates["s1_start"]) < 2:
        raise NotImplementedError(
            f"Recording {file_path.name} does not contain enough S1-start annotations "
            f"to define cycles."
        )

    min_cycle_samples = int(PREPROCESS["min_cycle_seconds"] * sampling_rate)
    max_cycle_samples = int(PREPROCESS["max_cycle_seconds"] * sampling_rate)

    metadata_rows: list[dict[str, Any]] = []
    feature_records: list[dict[str, Any]] = []
    waveform_records: list[dict[str, Any]] = []
    recording_exclusion_counter: Counter[str] = Counter()

    s1_starts = event_candidates["s1_start"]
    for cycle_index in range(len(s1_starts) - 1):
        s1_start_event = s1_starts[cycle_index]
        next_s1_start_event = s1_starts[cycle_index + 1]

        cycle_start_sample = int(s1_start_event["sample_index"])
        next_s1_start_sample = int(next_s1_start_event["sample_index"])
        cycle_start_row = int(s1_start_event["peak_row"])
        cycle_end_row_exclusive = int(next_s1_start_event["peak_row"])
        cycle_length_samples = next_s1_start_sample - cycle_start_sample
        cycle_length_in_bounds_flag = int(
            min_cycle_samples <= cycle_length_samples <= max_cycle_samples
        )

        sample_id = f"{identity['recording_id']}_cycle_{cycle_index:04d}"
        cycle_metadata: dict[str, Any] = {
            "sample_id": sample_id,
            "subject_id": identity["subject_id"],
            "recording_id": identity["recording_id"],
            "auscultation_site": identity["auscultation_site"],
            "source_file": file_path.name,
            "cycle_index": cycle_index,
            "cycle_start_sample": cycle_start_sample,
            "cycle_end_sample": next_s1_start_sample,
            "s1_start_sample": cycle_start_sample,
            "next_s1_start_sample": next_s1_start_sample,
            "original_cycle_length_samples": cycle_length_samples,
            "fixed_length": PREPROCESS["fixed_length"],
            "sampling_rate": sampling_rate,
            "qc_cycle_length_in_bounds_flag": cycle_length_in_bounds_flag,
            "valid_flag": False,
            "exclusion_reason": "",
            "feature_row_index": np.nan,
            "waveform_row_index": np.nan,
        }

        s1_end_candidates = events_between(
            event_candidates["s1_end"], cycle_start_sample, next_s1_start_sample
        )
        s2_start_candidates = events_between(
            event_candidates["s2_start"], cycle_start_sample, next_s1_start_sample
        )
        s2_end_candidates = events_between(
            event_candidates["s2_end"], cycle_start_sample, next_s1_start_sample
        )

        selected_events, exclusion_reason, diagnostics = best_valid_event_combination(
            cycle_start_sample=cycle_start_sample,
            next_s1_start_sample=next_s1_start_sample,
            s1_end_candidates=s1_end_candidates,
            s2_start_candidates=s2_start_candidates,
            s2_end_candidates=s2_end_candidates,
        )

        cycle_metadata.update(
            {
                "num_s1_end_candidates": diagnostics["num_s1_end_candidates"],
                "num_s2_start_candidates": diagnostics["num_s2_start_candidates"],
                "num_s2_end_candidates": diagnostics["num_s2_end_candidates"],
                "qc_selected_from_multiple_candidates_flag": diagnostics[
                    "selected_from_multiple_candidates_flag"
                ],
                "event_selection_strategy": diagnostics.get("event_selection_strategy", ""),
                "qc_event_ordering_valid_flag": int(exclusion_reason != "invalid_event_order"),
            }
        )

        if selected_events is None:
            cycle_metadata["exclusion_reason"] = exclusion_reason
            metadata_rows.append(cycle_metadata)
            recording_exclusion_counter[exclusion_reason] += 1
            continue

        s1_end_event = selected_events["s1_end"]
        s2_start_event = selected_events["s2_start"]
        s2_end_event = selected_events["s2_end"]

        cycle_metadata["s1_end_sample"] = int(s1_end_event["sample_index"])
        cycle_metadata["s2_start_sample"] = int(s2_start_event["sample_index"])
        cycle_metadata["s2_end_sample"] = int(s2_end_event["sample_index"])

        if (not cycle_length_in_bounds_flag) and QC["hard_exclude_out_of_bounds_cycles"]:
            cycle_metadata["exclusion_reason"] = "cycle_length_out_of_bounds"
            metadata_rows.append(cycle_metadata)
            recording_exclusion_counter["cycle_length_out_of_bounds"] += 1
            continue

        s1_end_row = int(s1_end_event["peak_row"])
        s2_start_row = int(s2_start_event["peak_row"])
        s2_end_row = int(s2_end_event["peak_row"])

        raw_cycle_waveform = filtered_amplitude[cycle_start_row:cycle_end_row_exclusive]
        normalized_cycle = normalize_cycle_waveform(
            raw_cycle_waveform,
            subtract_mean=FEATURE["subtract_cycle_mean_before_normalization"],
        )

        raw_segment_map = {
            "cycle": raw_cycle_waveform,
            "s1": filtered_amplitude[cycle_start_row:s1_end_row],
            "systole": filtered_amplitude[s1_end_row:s2_start_row],
            "s2": filtered_amplitude[s2_start_row:s2_end_row],
            "diastole": filtered_amplitude[s2_end_row:cycle_end_row_exclusive],
        }
        normalized_segment_map = {
            "cycle": normalized_cycle,
            "s1": normalized_cycle[: s1_end_row - cycle_start_row],
            "systole": normalized_cycle[s1_end_row - cycle_start_row : s2_start_row - cycle_start_row],
            "s2": normalized_cycle[s2_start_row - cycle_start_row : s2_end_row - cycle_start_row],
            "diastole": normalized_cycle[s2_end_row - cycle_start_row :],
        }

        if any(segment.size == 0 for segment in raw_segment_map.values()):
            cycle_metadata["exclusion_reason"] = "empty_segment_after_ordering"
            metadata_rows.append(cycle_metadata)
            recording_exclusion_counter["empty_segment_after_ordering"] += 1
            continue
        if any(segment.size == 0 for segment in normalized_segment_map.values()):
            cycle_metadata["exclusion_reason"] = "empty_segment_after_ordering"
            metadata_rows.append(cycle_metadata)
            recording_exclusion_counter["empty_segment_after_ordering"] += 1
            continue

        event_rows = {
            "s1_start": cycle_start_row,
            "s1_end": s1_end_row,
            "s2_start": s2_start_row,
            "s2_end": s2_end_row,
        }
        event_samples = {
            "s1_start": cycle_start_sample,
            "s1_end": int(s1_end_event["sample_index"]),
            "s2_start": int(s2_start_event["sample_index"]),
            "s2_end": int(s2_end_event["sample_index"]),
        }

        full_feature_row = build_full_feature_row(
            recording_table=recording_table,
            event_rows=event_rows,
            raw_segment_map=raw_segment_map,
            normalized_segment_map=normalized_segment_map,
            event_samples=event_samples,
            next_s1_start_sample=next_s1_start_sample,
            sampling_rate=sampling_rate,
            cycle_length_in_bounds_flag=cycle_length_in_bounds_flag,
            diagnostics=diagnostics,
        )
        feature_row = select_enabled_features(full_feature_row, feature_config=FEATURE)

        fixed_waveform = fixed_length_representation(
            normalized_cycle,
            fixed_length=PREPROCESS["fixed_length"],
        )

        cycle_metadata.update(
            {
                "time_cycle_duration_sec": full_feature_row["time_cycle_duration_sec"],
                "time_s1_duration_sec": full_feature_row["time_s1_duration_sec"],
                "time_systole_duration_sec": full_feature_row["time_systole_duration_sec"],
                "time_s2_duration_sec": full_feature_row["time_s2_duration_sec"],
                "time_diastole_duration_sec": full_feature_row["time_diastole_duration_sec"],
                "time_s1_to_s2_interval_sec": full_feature_row["time_s1_to_s2_interval_sec"],
                "time_s2_to_next_s1_interval_sec": full_feature_row["time_s2_to_next_s1_interval_sec"],
                "valid_flag": True,
                "exclusion_reason": "",
            }
        )

        metadata_rows.append(cycle_metadata)
        feature_records.append({"sample_id": sample_id, **feature_row})
        waveform_records.append({"sample_id": sample_id, "waveform": fixed_waveform})

    recording_summary = {
        "recording_id": identity["recording_id"],
        "subject_id": identity["subject_id"],
        "auscultation_site": identity["auscultation_site"],
        "source_file": file_path.name,
        "num_samples": int(len(recording_table)),
        "num_s1_start_events": int(len(event_candidates["s1_start"])),
        "num_candidate_cycles": int(max(len(event_candidates["s1_start"]) - 1, 0)),
        "num_valid_cycles": int(sum(row["valid_flag"] for row in metadata_rows)),
        "num_excluded_cycles": int(sum(not row["valid_flag"] for row in metadata_rows)),
        "exclusion_reasons": dict(recording_exclusion_counter),
    }

    return metadata_rows, feature_records, waveform_records, recording_summary


def export_preprocess_excel(
    preprocess_root: Path,
    summary: dict[str, Any],
    metadata_frame: pd.DataFrame,
    feature_matrix: np.ndarray,
    feature_names: list[str],
    feature_metadata_rows: list[dict[str, Any]],
    valid_metadata: pd.DataFrame,
    excluded_metadata: pd.DataFrame,
) -> Path:
    """Export preprocess artifacts to an Excel workbook for manual review."""
    overview_rows = [
        {"section": "run", "metric": "run_name", "value": summary["run_name"]},
        {"section": "data", "metric": "data_root", "value": summary["data_root"]},
        {"section": "data", "metric": "num_input_files", "value": summary["num_input_files"]},
        {"section": "cycles", "metric": "num_candidate_cycles", "value": summary["num_candidate_cycles"]},
        {"section": "cycles", "metric": "num_valid_cycles", "value": summary["num_valid_cycles"]},
        {"section": "cycles", "metric": "num_excluded_cycles", "value": summary["num_excluded_cycles"]},
        {"section": "features", "metric": "feature_rows", "value": summary["feature_shape"][0]},
        {"section": "features", "metric": "feature_dim", "value": summary["feature_shape"][1]},
        {"section": "features", "metric": "feature_group_mode", "value": summary["feature_selection"]["feature_group_mode"]},
        {"section": "waveforms", "metric": "waveform_rows", "value": summary["waveform_shape"][0]},
        {"section": "waveforms", "metric": "waveform_length", "value": summary["waveform_shape"][1]},
    ]
    overview_df = pd.DataFrame(overview_rows)

    feature_group_summary_df = pd.DataFrame(
        [
            {"feature_group": group_name, "feature_count": count}
            for group_name, count in summary["feature_group_counts"].items()
        ]
    )

    recording_summary_df = pd.DataFrame(summary["recordings"])
    if "exclusion_reasons" in recording_summary_df.columns:
        recording_summary_df["exclusion_reasons_json"] = recording_summary_df["exclusion_reasons"].apply(
            lambda value: json.dumps(value, ensure_ascii=False)
        )
        recording_summary_df = recording_summary_df.drop(columns=["exclusion_reasons"])

    feature_names_df = pd.DataFrame(feature_metadata_rows)

    if feature_names:
        feature_frame = pd.DataFrame(feature_matrix, columns=feature_names)
        feature_frame["feature_row_index"] = np.arange(len(feature_frame), dtype=int)
        valid_feature_view = valid_metadata.merge(feature_frame, on="feature_row_index", how="left")
    else:
        valid_feature_view = valid_metadata.copy()

    workbook_path = preprocess_root / EXCEL["filename"]
    sheets = {
        "Overview": overview_df,
        "Feature_Group_Summary": feature_group_summary_df,
        "Valid_Cycle_View": valid_feature_view,
        "Excluded_Cycles": excluded_metadata,
        "Recording_Summary": recording_summary_df,
        "Feature_Names": feature_names_df,
        "All_Metadata": metadata_frame,
    }
    return export_stage_workbook(
        workbook_path=workbook_path,
        sheets=sheets,
        freeze_panes=EXCEL["freeze_panes"],
        header_fill=EXCEL["header_fill"],
        header_font_color=EXCEL["header_font_color"],
        max_column_width=EXCEL["max_column_width"],
    )


def main() -> None:
    """Run the standalone preprocessing pipeline from raw Excel files."""
    set_random_seed(RANDOM_SEED)
    output_paths = ensure_output_directories(PATHS["output_root"], RUN_NAME)

    data_root = PATHS["data_root"]
    if not data_root.exists():
        raise FileNotFoundError(f"Data root does not exist: {data_root}")

    file_paths = sorted(data_root.glob(DATA["file_glob"]))
    file_paths = [
        file_path for file_path in file_paths if not file_path.name.startswith("~$")
    ]
    if not file_paths:
        raise FileNotFoundError(f"No input files found under: {data_root}")

    all_metadata_rows: list[dict[str, Any]] = []
    all_feature_records: list[dict[str, Any]] = []
    all_waveform_records: list[dict[str, Any]] = []
    recording_summaries: list[dict[str, Any]] = []

    for file_path in file_paths:
        logger.info("Processing recording: %s", file_path.name)
        metadata_rows, feature_records, waveform_records, recording_summary = process_recording(
            file_path=file_path,
            sampling_rate=DATA["sampling_rate"],
        )
        all_metadata_rows.extend(metadata_rows)
        all_feature_records.extend(feature_records)
        all_waveform_records.extend(waveform_records)
        recording_summaries.append(recording_summary)

    metadata_frame = pd.DataFrame(all_metadata_rows)
    metadata_frame = metadata_frame.sort_values(
        by=["recording_id", "cycle_index"], kind="stable"
    ).reset_index(drop=True)

    valid_metadata = metadata_frame.loc[metadata_frame["valid_flag"] == True].copy()
    excluded_metadata = metadata_frame.loc[metadata_frame["valid_flag"] == False].copy()

    if not valid_metadata.empty:
        valid_metadata["feature_row_index"] = np.arange(len(valid_metadata), dtype=int)
        valid_metadata["waveform_row_index"] = np.arange(len(valid_metadata), dtype=int)
        metadata_frame.loc[valid_metadata.index, "feature_row_index"] = valid_metadata["feature_row_index"].to_numpy()
        metadata_frame.loc[valid_metadata.index, "waveform_row_index"] = valid_metadata["waveform_row_index"].to_numpy()

    if all_feature_records:
        feature_frame = pd.DataFrame(all_feature_records)
        feature_frame = feature_frame.merge(
            valid_metadata[["sample_id", "feature_row_index"]],
            on="sample_id",
            how="inner",
        )
        feature_frame = feature_frame.sort_values("feature_row_index", kind="stable").reset_index(drop=True)
        feature_names = [column for column in feature_frame.columns if column not in {"sample_id", "feature_row_index"}]
        feature_matrix = feature_frame[feature_names].to_numpy(dtype=np.float32)

        waveform_frame = pd.DataFrame(
            {
                "sample_id": [record["sample_id"] for record in all_waveform_records],
                "waveform": [record["waveform"] for record in all_waveform_records],
            }
        )
        waveform_frame = waveform_frame.merge(
            valid_metadata[["sample_id", "waveform_row_index"]],
            on="sample_id",
            how="inner",
        )
        waveform_frame = waveform_frame.sort_values("waveform_row_index", kind="stable").reset_index(drop=True)
        waveform_matrix = np.stack(waveform_frame["waveform"].tolist()).astype(np.float32)
    else:
        feature_names = []
        feature_matrix = np.empty((0, 0), dtype=np.float32)
        waveform_matrix = np.empty((0, PREPROCESS["fixed_length"]), dtype=np.float32)

    feature_metadata_rows = build_feature_metadata(feature_names)
    feature_group_counts = Counter(
        row["feature_group"] for row in feature_metadata_rows
    )

    preprocess_root = output_paths["preprocess_root"]
    np.save(preprocess_root / "cycle_features.npy", feature_matrix)
    np.save(preprocess_root / "cycle_waveforms.npy", waveform_matrix)
    metadata_frame.to_csv(preprocess_root / "cycle_metadata.csv", index=False)

    with open(preprocess_root / "feature_names.json", "w", encoding="utf-8") as file:
        json.dump(feature_names, file, indent=2, ensure_ascii=False)
    with open(preprocess_root / "feature_metadata.json", "w", encoding="utf-8") as file:
        json.dump(feature_metadata_rows, file, indent=2, ensure_ascii=False)

    exclusion_counts = excluded_metadata["exclusion_reason"].value_counts().to_dict()
    qc_selected_from_multiple_summary = summarize_boolean_array(
        valid_metadata["qc_selected_from_multiple_candidates_flag"].to_numpy(dtype=np.float32) > 0
        if "qc_selected_from_multiple_candidates_flag" in valid_metadata.columns
        else np.array([], dtype=bool)
    )
    qc_cycle_bounds_summary = summarize_boolean_array(
        valid_metadata["qc_cycle_length_in_bounds_flag"].to_numpy(dtype=np.float32) > 0
        if "qc_cycle_length_in_bounds_flag" in valid_metadata.columns
        else np.array([], dtype=bool)
    )

    summary = {
        "run_name": RUN_NAME,
        "data_root": str(data_root),
        "num_input_files": len(file_paths),
        "sampling_rate": DATA["sampling_rate"],
        "fixed_length": PREPROCESS["fixed_length"],
        "feature_shape": list(feature_matrix.shape),
        "waveform_shape": list(waveform_matrix.shape),
        "num_candidate_cycles": int(len(metadata_frame)),
        "num_valid_cycles": int(len(valid_metadata)),
        "num_excluded_cycles": int(len(excluded_metadata)),
        "valid_cycle_duration_sec": summarize_array(
            valid_metadata["time_cycle_duration_sec"].to_numpy(dtype=np.float32)
            if "time_cycle_duration_sec" in valid_metadata.columns and not valid_metadata.empty
            else np.array([], dtype=np.float32)
        ),
        "feature_group_counts": dict(sorted(feature_group_counts.items())),
        "feature_selection": {
            "feature_group_mode": FEATURE["feature_group_mode"],
            "selected_prefixes": sorted(select_feature_prefixes(FEATURE)),
            "enable_rs_features": FEATURE["enable_rs_features"],
            "enable_time_features": FEATURE["enable_time_features"],
            "enable_amplitude_features": FEATURE["enable_amplitude_features"],
            "enable_raw_amplitude_features": FEATURE["enable_raw_amplitude_features"],
            "enable_normalized_amplitude_features": FEATURE["enable_normalized_amplitude_features"],
            "enable_ratio_features": FEATURE["enable_ratio_features"],
            "enable_qc_features": FEATURE["enable_qc_features"],
        },
        "qc_summaries": {
            "selected_from_multiple_candidates": qc_selected_from_multiple_summary,
            "cycle_length_in_bounds": qc_cycle_bounds_summary,
        },
        "exclusion_reasons": exclusion_counts,
        "recordings": recording_summaries,
    }

    with open(preprocess_root / "preprocess_summary.json", "w", encoding="utf-8") as file:
        json.dump(summary, file, indent=2, ensure_ascii=False)

    save_qc_plots(
        metadata_frame=metadata_frame,
        waveform_matrix=waveform_matrix,
        qc_root=output_paths["qc_root"],
        fixed_length=PREPROCESS["fixed_length"],
        random_seed=RANDOM_SEED,
    )

    excel_path = None
    if EXCEL["export_enabled"]:
        excel_path = export_preprocess_excel(
            preprocess_root=preprocess_root,
            summary=summary,
            metadata_frame=metadata_frame,
            feature_matrix=feature_matrix,
            feature_names=feature_names,
            feature_metadata_rows=feature_metadata_rows,
            valid_metadata=valid_metadata,
            excluded_metadata=excluded_metadata,
        )

    logger.info("Saved preprocess outputs to: %s", preprocess_root)
    if excel_path is not None:
        logger.info("Saved preprocess Excel export to: %s", excel_path)
    logger.info("Feature matrix shape: %s", feature_matrix.shape)
    logger.info("Waveform matrix shape: %s", waveform_matrix.shape)
    logger.info("Metadata rows: %s", len(metadata_frame))
    logger.info("Valid cycles: %s", len(valid_metadata))
    logger.info("Excluded cycles: %s", len(excluded_metadata))


if __name__ == "__main__":
    main()
